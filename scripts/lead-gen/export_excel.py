"""
Professional Excel Export for Muqawil Contractor Leads (All Regions)
"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_FILE = Path("data/muqawil_all_regions.json")
OUTPUT_DIR = Path("data")

with open(DATA_FILE, encoding="utf-8") as f:
    all_data = json.load(f)

def filter_has_email(d):
    return [r for r in d if r.get("email") and len(r["email"]) > 3]

def filter_has_phone(d):
    return [r for r in d if r.get("phone") and len(r["phone"]) > 3]

def filter_region(d, region_name):
    return [r for r in d if r.get("region_name") == region_name or r.get("region_name") == region_name]

# Normalize
for r in all_data:
    city = r.get("cityName") or r.get("city") or ""
    r["city_display"] = city
    r["region_display"] = r.get("region_name", "")
    if r.get("companySize"):
        size_map = {"small": "منشأة صغيرة", "medium": "منشأة متوسطة", "large": "منشأة كبيرة", "verysmall": "منشأة متناهية الصغر"}
        r["companySize"] = size_map.get(r["companySize"], r["companySize"])
    if r.get("phone"):
        r["phone"] = re.sub(r"[^\d+]", "", r["phone"])

all_data = [r for r in all_data if r.get("companyName")]
all_data.sort(key=lambda r: (r.get("region_display", ""), r.get("city_display", ""), r.get("companyName", "")))

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="center")
EMAIL_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

COLUMNS = [
    ("#", 5),
    ("المنطقة", 16),
    ("المدينة", 15),
    ("اسم الشركة", 35),
    ("رقم العضوية", 18),
    ("الهاتف", 18),
    ("البريد الإلكتروني", 30),
    ("حجم المنشأة", 18),
]

def write_sheet(ws, data):
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, record in enumerate(data, 2):
        is_alt = row_idx % 2 == 0
        row_fill = ALT_FILL if is_alt else PatternFill()

        vals = [
            row_idx - 1,
            record.get("region_display", ""),
            record.get("city_display", ""),
            record.get("companyName", ""),
            record.get("membershipNo", ""),
            record.get("phone", ""),
            record.get("email", ""),
            record.get("companySize", ""),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = THIN_BORDER

            if col_idx == 1:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = DATA_ALIGN
            elif col_idx == 3:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = DATA_ALIGN
            elif col_idx == 5:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val:
                    cell.number_format = '@'
            elif col_idx == 7:
                if val:
                    cell.font = EMAIL_FONT
                    cell.hyperlink = f"mailto:{val}"
                else:
                    cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
            elif col_idx == 8:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

    n = len(data)
    ws.auto_filter.ref = f"A1:H{n + 1}" if n else "A1:H1"
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# ── Create workbook ──
wb = Workbook()

# Sheet 1: All contractors
ws = wb.active
ws.title = "الكل"
write_sheet(ws, all_data)

# Sheet 2: With Email
ws2 = wb.create_sheet("بإيميل")
write_sheet(ws2, filter_has_email(all_data))

# Sheet 3-5: Major regions
REGION_SHEETS = [
    ("الرياض", lambda: filter_region(all_data, "الرياض")),
    ("مكة المكرمة", lambda: filter_region(all_data, "مكة المكرمة")),
    ("الشرقية", lambda: filter_region(all_data, "الشرقية")),
]
for sheet_name, filter_fn in REGION_SHEETS:
    wsn = wb.create_sheet(sheet_name)
    write_sheet(wsn, filter_fn())

# Sheet: Statistics
ws_stat = wb.create_sheet("إحصائيات")

# Region stats
region_counts = {}
for r in all_data:
    reg = r.get("region_display") or "غير محدد"
    if reg not in region_counts:
        region_counts[reg] = {"total": 0, "with_email": 0, "with_phone": 0}
    region_counts[reg]["total"] += 1
    if r.get("email"): region_counts[reg]["with_email"] += 1
    if r.get("phone"): region_counts[reg]["with_phone"] += 1

sorted_regions = sorted(region_counts.items(), key=lambda x: x[1]["total"], reverse=True)

stat_headers = ["المنطقة", "الإجمالي", "بإيميل", "بهاتف"]
stat_widths = [20, 14, 12, 12]

for col_idx, (h, w) in enumerate(zip(stat_headers, stat_widths), 1):
    cell = ws_stat.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER
    ws_stat.column_dimensions[get_column_letter(col_idx)].width = w

for row_idx, (region, counts) in enumerate(sorted_regions, 2):
    is_alt = row_idx % 2 == 0
    fill = ALT_FILL if is_alt else PatternFill()
    cell = ws_stat.cell(row=row_idx, column=1, value=region)
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.alignment = DATA_ALIGN
    cell.fill = fill
    cell.border = THIN_BORDER
    for col_idx, key in enumerate(["total", "with_email", "with_phone"], 2):
        cell = ws_stat.cell(row=row_idx, column=col_idx, value=counts[key])
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = THIN_BORDER

# Totals row
total_row = len(sorted_regions) + 2
totals = {"total": 0, "with_email": 0, "with_phone": 0}
for _, c in region_counts.items():
    for k in totals: totals[k] += c[k]

totals_data = [("المنطقة", "المجموع الكلي"), ("total", totals["total"]), ("with_email", totals["with_email"]), ("with_phone", totals["with_phone"])]
col_keys = ["", "total", "with_email", "with_phone"]
for label, val in totals_data:
    ci = 1 if label == "المنطقة" else col_keys.index(label) + 1
    cell = ws_stat.cell(row=total_row, column=ci, value=val)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = DATA_ALIGN if ci == 1 else Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

ws_stat.freeze_panes = "A2"

# ── Save ──
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
output_path = OUTPUT_DIR / "muqawil_contractors_professional.xlsx"
wb.save(output_path)

email_count = len(filter_has_email(all_data))
phone_count = len(filter_has_phone(all_data))

print(f"✓ ملف Excel: {output_path}")
print(f"  - الإجمالي: {len(all_data)}")
print(f"  - بإيميل: {email_count}")
print(f"  - بهاتف: {phone_count}")
print(f"  - المناطق: {len(region_counts)}")
