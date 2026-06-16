import json, os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DATA_DIR = 'data'
TENDERS_FILE = os.path.join(DATA_DIR, 'etimad_all_tenders.json')
CONTRACTORS_FILE = os.path.join(DATA_DIR, 'muqawil_all_regions.json')
OUTPUT_FILE = os.path.join(DATA_DIR, 'منتجع_المنافسات_والمقاولين.xlsx')

# ── Load data ──
print('Loading data...')
with open(TENDERS_FILE, 'r', encoding='utf-8') as f:
    tenders = json.load(f)
with open(CONTRACTORS_FILE, 'r', encoding='utf-8') as f:
    contractors = json.load(f)

print(f'Tenders: {len(tenders):,}')
print(f'Contractors: {len(contractors):,}')

def clean_str(s):
    """Remove illegal XML characters for openpyxl"""
    if not s:
        return ''
    if not isinstance(s, str):
        s = str(s)
    # Remove control characters except \t, \n, \r
    import re
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)

# ── Construction-related activities ──
CONSTRUCTION_KEYWORDS = [
    'مقاولات', 'تشييد', 'بناء', 'إنشاء', 'هدم', 'ترميم', 'صيانة',
    'تشغيل', 'نظافة', 'كهرباء', 'سباكة', 'تكييف', 'مياه',
    'طرق', 'بنية تحتية', ' infrastruct', 'أسفلت', 'خرسانة',
    'حديد', 'دهان', 'عزل', 'زجاج', 'ألمنيوم', 'سيراميك',
    'مواد بناء', 'معدات', 'آليات', 'رافعة',
    'حفر', 'ردم', 'تسوية', 'تأثيث', 'أثاث'
]

def is_construction_tender(t):
    name = (t.get('tenderName') or '') + ' ' + (t.get('tenderActivityName') or '') + ' ' + (t.get('tenderTypeName') or '')
    name_lower = name.lower()
    for kw in CONSTRUCTION_KEYWORDS:
        if kw in name_lower or kw in name:
            return True
    return False

# ── Filter construction tenders ──
construction_tenders = [t for t in tenders if is_construction_tender(t)]
print(f'Construction tenders: {len(construction_tenders):,}')

# ── Expiring soon (7 days) ──
now = datetime.now()
week_from_now = now + timedelta(days=7)
expiring_soon = []
for t in tenders:
    d_str = t.get('lastOfferPresentationDate') or ''
    if d_str and d_str != '0001-01-01T00:00:00':
        try:
            d = datetime.fromisoformat(d_str.replace('Z', '+00:00'))
            if now < d < week_from_now:
                expiring_soon.append(t)
        except:
            pass
print(f'Expiring within 7 days: {len(expiring_soon):,}')

# ── Contractors stats ──
with_email = [c for c in contractors if c.get('email')]
with_phone = [c for c in contractors if c.get('phone')]
print(f'Contractors with email: {len(with_email):,}')
print(f'Contractors with phone: {len(with_phone):,}')

# ── Region distribution ──
region_counts = {}
for c in contractors:
    r = c.get('region_name') or 'غير محدد'
    region_counts[r] = region_counts.get(r, 0) + 1

# ── Activity distribution ──
activity_counts = {}
for t in tenders:
    a = t.get('tenderActivityName') or 'غير محدد'
    activity_counts[a] = activity_counts.get(a, 0) + 1

# ── Agency distribution ──
agency_counts = {}
for t in tenders:
    a = t.get('agencyName') or 'غير محدد'
    agency_counts[a] = agency_counts.get(a, 0) + 1

# ── Status distribution ──
status_counts = {}
status_map = {2: 'نشطة', 3: 'فتح العروض', 4: 'فحص العروض', 5: 'الترسية', 6: 'تم الترسية', 8: 'منتهية'}
for t in tenders:
    s = status_map.get(t.get('tenderStatusId'), str(t.get('tenderStatusId', '?')))
    status_counts[s] = status_counts.get(s, 0) + 1

# ═══════════════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════════════

wb = Workbook()

# Colors
DARK_BLUE = '1F3864'
MED_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
GOLD = 'FFD700'
WHITE = 'FFFFFF'
LIGHT_GRAY = 'F2F2F2'
GREEN = '28A745'
RED = 'DC3545'

hdr_font = Font(name='Arial', bold=True, color=WHITE, size=11)
hdr_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type='solid')
title_font = Font(name='Arial', bold=True, color=DARK_BLUE, size=16)
subtitle_font = Font(name='Arial', bold=True, color=MED_BLUE, size=12)
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
link_font = Font(name='Arial', size=10, color='0563C1', underline='single')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill

def add_title(ws, title, subtitle='', start_row=1):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal='right', vertical='center')
    if subtitle:
        ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=10)
        cell2 = ws.cell(row=start_row+1, column=1, value=subtitle)
        cell2.font = subtitle_font
        cell2.alignment = Alignment(horizontal='right', vertical='center')
    return start_row + 3

# ──────── SHEET 1: Cover ────────
ws_cover = wb.active
ws_cover.title = 'غلاف'
ws_cover.sheet_properties.tabColor = GOLD

ws_cover.merge_cells('A1:J1')
c = ws_cover.cell(row=1, column=1, value='منتج المنافسات الحكومية والمقاولين')
c.font = Font(name='Arial', bold=True, size=24, color=DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_cover.row_dimensions[1].height = 50

ws_cover.merge_cells('A3:J3')
c = ws_cover.cell(row=3, column=1, value='Government Tenders & Contractors Database')
c.font = Font(name='Arial', size=14, color=MED_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')

ws_cover.merge_cells('A5:J5')
c = ws_cover.cell(row=5, column=1, value=f'تاريخ التصدير: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
c.font = Font(name='Arial', size=11, color='666666')
c.alignment = Alignment(horizontal='center')

info = [
    ('', ''),
    ('المحتويات', 'الوصف'),
    ('كل المنافسات', f'{len(tenders):,} منافسة حكومية من منصة اعتماد'),
    ('منافسات التشييد', f'{len(construction_tenders):,} منافسة مرتبطة بقطاع التشييد والمقاولات'),
    ('منافسات تنتهي قريباً', f'{len(expiring_soon):,} منافسة تنتهي خلال 7 أيام'),
    ('دليل المقاولين', f'{len(contractors):,} مقاول عبر 13 منطقة سعودية'),
    ('مقاولين بإيميل', f'{len(with_email):,} مقاول مع بريد إلكتروني'),
    ('مقاولين بجوال', f'{len(with_phone):,} مقاول مع رقم جوال'),
    ('', ''),
    ('مصادر البيانات:', ''),
    ('المنافسات', 'منصة اعتماد (tenders.etimad.sa) — 8,160 منافسة'),
    ('المقاولين', 'موقع Muqawil.org — 13,514 مقاول'),
]

for i, (k, v) in enumerate(info):
    r = 8 + i
    ws_cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws_cover.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
    ck = ws_cover.cell(row=r, column=2, value=k)
    cv = ws_cover.cell(row=r, column=6, value=v)
    if k in ('المحتويات', 'مصادر البيانات:'):
        ck.font = Font(name='Arial', bold=True, size=12, color=DARK_BLUE)
    elif k:
        ck.font = Font(name='Arial', bold=True, size=11)
        cv.font = Font(name='Arial', size=11)
    else:
        ck.font = Font(name='Arial', size=10)

ws_cover.column_dimensions['A'].width = 3
for col in ['B','C','D','E']: ws_cover.column_dimensions[col].width = 18
for col in ['F','G','H','I','J']: ws_cover.column_dimensions[col].width = 18

# ──────── SHEET 2: All Tenders ────────
ws_all = wb.create_sheet('كل المنافسات')
ws_all.sheet_properties.tabColor = MED_BLUE

tender_headers = [
    'رقم المنافسة', 'اسم المنافسة', 'الجهة الحكومية', 'الفرع',
    'النشاط', 'نوع المنافسة', 'تاريخ النشر', 'آخر موعد للتقديم',
    'تاريخ فتح العروض', 'سعر الكراسة', 'الحالة', 'الأيام المتبقية'
]
for c, h in enumerate(tender_headers, 1):
    ws_all.cell(row=1, column=c, value=h)
style_header_row(ws_all, 1, len(tender_headers))

for i, t in enumerate(tenders, 2):
    ws_all.cell(row=i, column=1, value=clean_str(t.get('tenderNumber', '')))
    ws_all.cell(row=i, column=2, value=clean_str(t.get('tenderName', '')))
    ws_all.cell(row=i, column=3, value=clean_str(t.get('agencyName', '')))
    ws_all.cell(row=i, column=4, value=clean_str(t.get('branchName', '')))
    ws_all.cell(row=i, column=5, value=clean_str(t.get('tenderActivityName', '')))
    ws_all.cell(row=i, column=6, value=clean_str(t.get('tenderTypeName', '')))
    ws_all.cell(row=i, column=7, value=clean_str(t.get('submitionDate', '')[:10] if t.get('submitionDate') else ''))
    ws_all.cell(row=i, column=8, value=clean_str(t.get('lastOfferPresentationDate', '')[:10] if t.get('lastOfferPresentationDate') else ''))
    ws_all.cell(row=i, column=9, value=clean_str(t.get('offersOpeningDate', '')[:10] if t.get('offersOpeningDate') else ''))
    ws_all.cell(row=i, column=10, value=clean_str(t.get('condetionalBookletPrice', '')))
    ws_all.cell(row=i, column=11, value=clean_str(status_map.get(t.get('tenderStatusId'), '')))
    ws_all.cell(row=i, column=12, value=clean_str(t.get('remainingDays', '')))

style_data_rows(ws_all, 2, len(tenders) + 1, len(tender_headers))

# Column widths
widths_all = [15, 45, 30, 25, 22, 15, 14, 14, 14, 12, 14, 12]
for i, w in enumerate(widths_all, 1):
    ws_all.column_dimensions[get_column_letter(i)].width = w

ws_all.auto_filter.ref = f'A1:{get_column_letter(len(tender_headers))}{len(tenders)+1}'
ws_all.freeze_panes = 'A2'

# ──────── SHEET 3: Construction Tenders ────────
ws_con = wb.create_sheet('منافسات التشييد')
ws_con.sheet_properties.tabColor = '28A745'

for c, h in enumerate(tender_headers, 1):
    ws_con.cell(row=1, column=c, value=h)
style_header_row(ws_con, 1, len(tender_headers))

for i, t in enumerate(construction_tenders, 2):
    ws_con.cell(row=i, column=1, value=clean_str(t.get('tenderNumber', '')))
    ws_con.cell(row=i, column=2, value=clean_str(t.get('tenderName', '')))
    ws_con.cell(row=i, column=3, value=clean_str(t.get('agencyName', '')))
    ws_con.cell(row=i, column=4, value=clean_str(t.get('branchName', '')))
    ws_con.cell(row=i, column=5, value=clean_str(t.get('tenderActivityName', '')))
    ws_con.cell(row=i, column=6, value=clean_str(t.get('tenderTypeName', '')))
    ws_con.cell(row=i, column=7, value=clean_str(t.get('submitionDate', '')[:10] if t.get('submitionDate') else ''))
    ws_con.cell(row=i, column=8, value=clean_str(t.get('lastOfferPresentationDate', '')[:10] if t.get('lastOfferPresentationDate') else ''))
    ws_con.cell(row=i, column=9, value=clean_str(t.get('offersOpeningDate', '')[:10] if t.get('offersOpeningDate') else ''))
    ws_con.cell(row=i, column=10, value=clean_str(t.get('condetionalBookletPrice', '')))
    ws_con.cell(row=i, column=11, value=clean_str(status_map.get(t.get('tenderStatusId'), '')))
    ws_con.cell(row=i, column=12, value=clean_str(t.get('remainingDays', '')))

style_data_rows(ws_con, 2, len(construction_tenders) + 1, len(tender_headers))
for i, w in enumerate(widths_all, 1):
    ws_con.column_dimensions[get_column_letter(i)].width = w
ws_con.auto_filter.ref = f'A1:{get_column_letter(len(tender_headers))}{len(construction_tenders)+1}'
ws_con.freeze_panes = 'A2'

# ──────── SHEET 4: Expiring Soon ────────
ws_exp = wb.create_sheet('تنتهي قريباً')
ws_exp.sheet_properties.tabColor = RED

for c, h in enumerate(tender_headers, 1):
    ws_exp.cell(row=1, column=c, value=h)
style_header_row(ws_exp, 1, len(tender_headers))

for i, t in enumerate(expiring_soon, 2):
    ws_exp.cell(row=i, column=1, value=clean_str(t.get('tenderNumber', '')))
    ws_exp.cell(row=i, column=2, value=clean_str(t.get('tenderName', '')))
    ws_exp.cell(row=i, column=3, value=clean_str(t.get('agencyName', '')))
    ws_exp.cell(row=i, column=4, value=clean_str(t.get('branchName', '')))
    ws_exp.cell(row=i, column=5, value=clean_str(t.get('tenderActivityName', '')))
    ws_exp.cell(row=i, column=6, value=clean_str(t.get('tenderTypeName', '')))
    ws_exp.cell(row=i, column=7, value=clean_str(t.get('submitionDate', '')[:10] if t.get('submitionDate') else ''))
    ws_exp.cell(row=i, column=8, value=clean_str(t.get('lastOfferPresentationDate', '')[:10] if t.get('lastOfferPresentationDate') else ''))
    ws_exp.cell(row=i, column=9, value=clean_str(t.get('offersOpeningDate', '')[:10] if t.get('offersOpeningDate') else ''))
    ws_exp.cell(row=i, column=10, value=clean_str(t.get('condetionalBookletPrice', '')))
    ws_exp.cell(row=i, column=11, value=clean_str(status_map.get(t.get('tenderStatusId'), '')))
    ws_exp.cell(row=i, column=12, value=clean_str(t.get('remainingDays', '')))

style_data_rows(ws_exp, 2, len(expiring_soon) + 1, len(tender_headers))
for i, w in enumerate(widths_all, 1):
    ws_exp.column_dimensions[get_column_letter(i)].width = w
ws_exp.auto_filter.ref = f'A1:{get_column_letter(len(tender_headers))}{len(expiring_soon)+1}'
ws_exp.freeze_panes = 'A2'

# ──────── SHEET 5: Contractors ────────
ws_conts = wb.create_sheet('دليل المقاولين')
ws_conts.sheet_properties.tabColor = 'FF8C00'

cont_headers = ['اسم الشركة', 'رقم العضوية', 'المنطقة', 'المدينة', 'الجوال', 'البريد الإلكتروني', 'حجم الشركة', 'المصدر']
for c, h in enumerate(cont_headers, 1):
    ws_conts.cell(row=1, column=c, value=h)
style_header_row(ws_conts, 1, len(cont_headers))

for i, cont in enumerate(contractors, 2):
    ws_conts.cell(row=i, column=1, value=clean_str(cont.get('companyName', '')))
    ws_conts.cell(row=i, column=2, value=clean_str(cont.get('membershipNo', '')))
    ws_conts.cell(row=i, column=3, value=clean_str(cont.get('region_name', '')))
    ws_conts.cell(row=i, column=4, value=clean_str(cont.get('cityName', '')))
    ws_conts.cell(row=i, column=5, value=clean_str(cont.get('phone', '')))
    ws_conts.cell(row=i, column=6, value=clean_str(cont.get('email', '')))
    ws_conts.cell(row=i, column=7, value=clean_str(cont.get('companySize', '')))
    ws_conts.cell(row=i, column=8, value=clean_str(cont.get('source', '')))

style_data_rows(ws_conts, 2, len(contractors) + 1, len(cont_headers))
widths_cont = [35, 15, 15, 15, 15, 25, 12, 10]
for i, w in enumerate(widths_cont, 1):
    ws_conts.column_dimensions[get_column_letter(i)].width = w
ws_conts.auto_filter.ref = f'A1:{get_column_letter(len(cont_headers))}{len(contractors)+1}'
ws_conts.freeze_panes = 'A2'

# ──────── SHEET 6: Contractors with Email ────────
ws_email = wb.create_sheet('مقاولين بإيميل')
ws_email.sheet_properties.tabColor = '17A2B8'

for c, h in enumerate(cont_headers, 1):
    ws_email.cell(row=1, column=c, value=h)
style_header_row(ws_email, 1, len(cont_headers))

for i, cont in enumerate(with_email, 2):
    ws_email.cell(row=i, column=1, value=clean_str(cont.get('companyName', '')))
    ws_email.cell(row=i, column=2, value=clean_str(cont.get('membershipNo', '')))
    ws_email.cell(row=i, column=3, value=clean_str(cont.get('region_name', '')))
    ws_email.cell(row=i, column=4, value=clean_str(cont.get('cityName', '')))
    ws_email.cell(row=i, column=5, value=clean_str(cont.get('phone', '')))
    ws_email.cell(row=i, column=6, value=clean_str(cont.get('email', '')))
    ws_email.cell(row=i, column=7, value=clean_str(cont.get('companySize', '')))
    ws_email.cell(row=i, column=8, value=clean_str(cont.get('source', '')))

style_data_rows(ws_email, 2, len(with_email) + 1, len(cont_headers))
for i, w in enumerate(widths_cont, 1):
    ws_email.column_dimensions[get_column_letter(i)].width = w
ws_email.auto_filter.ref = f'A1:{get_column_letter(len(cont_headers))}{len(with_email)+1}'
ws_email.freeze_panes = 'A2'

# ──────── SHEET 7: Statistics ────────
ws_stats = wb.create_sheet('إحصائيات')
ws_stats.sheet_properties.tabColor = '6F42C1'

add_title(ws_stats, 'إحصائيات المنافسات والمقاولين', f'آخر تحديث: {datetime.now().strftime("%Y-%m-%d %H:%M")}')

# Section 1: Overview
r = 5
ws_stats.merge_cells(f'A{r}:E{r}')
ws_stats.cell(row=r, column=1, value='نظرة عامة').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1

overview = [
    ('إجمالي المنافسات', f'{len(tenders):,}'),
    ('منافسات التشييد والمقاولات', f'{len(construction_tenders):,}'),
    ('منافسات تنتهي خلال 7 أيام', f'{len(expiring_soon):,}'),
    ('إجمالي المقاولين', f'{len(contractors):,}'),
    ('مقاولين مع بريد إلكتروني', f'{len(with_email):,}'),
    ('مقاولين مع رقم جوال', f'{len(with_phone):,}'),
    ('عدد المناطق', f'{len(region_counts)}'),
]

for k, v in overview:
    ws_stats.merge_cells(f'A{r}:B{r}')
    ws_stats.merge_cells(f'C{r}:E{r}')
    ws_stats.cell(row=r, column=1, value=k).font = bold_font
    ws_stats.cell(row=r, column=3, value=v).font = Font(name='Arial', size=12, color=MED_BLUE)
    ws_stats.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws_stats.cell(row=r, column=1).border = thin_border
    ws_stats.cell(row=r, column=3).border = thin_border
    ws_stats.cell(row=r, column=3).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    r += 1

# Section 2: Regions
r += 1
ws_stats.merge_cells(f'A{r}:E{r}')
ws_stats.cell(row=r, column=1, value='توزيع المقاولين حسب المنطقة').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1

ws_stats.merge_cells(f'A{r}:B{r}')
ws_stats.cell(row=r, column=1, value='المنطقة').font = hdr_font
ws_stats.cell(row=r, column=1).fill = hdr_fill
ws_stats.merge_cells(f'C{r}:E{r}')
ws_stats.cell(row=r, column=3, value='عدد المقاولين').font = hdr_font
ws_stats.cell(row=r, column=3).fill = hdr_fill
ws_stats.cell(row=r, column=1).border = thin_border
ws_stats.cell(row=r, column=3).border = thin_border
r += 1

sorted_regions = sorted(region_counts.items(), key=lambda x: x[1], reverse=True)
for region, count in sorted_regions:
    ws_stats.merge_cells(f'A{r}:B{r}')
    ws_stats.cell(row=r, column=1, value=clean_str(region)).font = normal_font
    ws_stats.merge_cells(f'C{r}:E{r}')
    ws_stats.cell(row=r, column=3, value=f'{count:,}').font = normal_font
    ws_stats.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws_stats.cell(row=r, column=1).border = thin_border
    ws_stats.cell(row=r, column=3).border = thin_border
    if (r - 7) % 2 == 0:
        ws_stats.cell(row=r, column=1).fill = alt_fill
        ws_stats.cell(row=r, column=3).fill = alt_fill
    r += 1

# Section 3: Top Agencies
r += 1
ws_stats.merge_cells(f'A{r}:E{r}')
ws_stats.cell(row=r, column=1, value='أكبر الجهات الحكومية طرحاً للمنافسات').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1

ws_stats.merge_cells(f'A{r}:B{r}')
ws_stats.cell(row=r, column=1, value='الجهة').font = hdr_font
ws_stats.cell(row=r, column=1).fill = hdr_fill
ws_stats.merge_cells(f'C{r}:E{r}')
ws_stats.cell(row=r, column=3, value='عدد المنافسات').font = hdr_font
ws_stats.cell(row=r, column=3).fill = hdr_fill
ws_stats.cell(row=r, column=1).border = thin_border
ws_stats.cell(row=r, column=3).border = thin_border
r += 1

sorted_agencies = sorted(agency_counts.items(), key=lambda x: x[1], reverse=True)[:30]
for agency, count in sorted_agencies:
    ws_stats.merge_cells(f'A{r}:B{r}')
    ws_stats.cell(row=r, column=1, value=clean_str(agency)).font = normal_font
    ws_stats.merge_cells(f'C{r}:E{r}')
    ws_stats.cell(row=r, column=3, value=f'{count:,}').font = normal_font
    ws_stats.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws_stats.cell(row=r, column=1).border = thin_border
    ws_stats.cell(row=r, column=3).border = thin_border
    if (r - (r - 15 if r > 15 else 7)) % 2 == 0:
        ws_stats.cell(row=r, column=1).fill = alt_fill
        ws_stats.cell(row=r, column=3).fill = alt_fill
    r += 1

# Section 4: Top Activities
r += 1
ws_stats.merge_cells(f'A{r}:E{r}')
ws_stats.cell(row=r, column=1, value='أكثر الأنشطة طلباً').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1

ws_stats.merge_cells(f'A{r}:B{r}')
ws_stats.cell(row=r, column=1, value='النشاط').font = hdr_font
ws_stats.cell(row=r, column=1).fill = hdr_fill
ws_stats.merge_cells(f'C{r}:E{r}')
ws_stats.cell(row=r, column=3, value='عدد المنافسات').font = hdr_font
ws_stats.cell(row=r, column=3).fill = hdr_fill
ws_stats.cell(row=r, column=1).border = thin_border
ws_stats.cell(row=r, column=3).border = thin_border
r += 1

sorted_activities = sorted(activity_counts.items(), key=lambda x: x[1], reverse=True)[:20]
for activity, count in sorted_activities:
    ws_stats.merge_cells(f'A{r}:B{r}')
    ws_stats.cell(row=r, column=1, value=clean_str(activity.strip())).font = normal_font
    ws_stats.merge_cells(f'C{r}:E{r}')
    ws_stats.cell(row=r, column=3, value=f'{count:,}').font = normal_font
    ws_stats.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws_stats.cell(row=r, column=1).border = thin_border
    ws_stats.cell(row=r, column=3).border = thin_border
    if (r - 7) % 2 == 0:
        ws_stats.cell(row=r, column=1).fill = alt_fill
        ws_stats.cell(row=r, column=3).fill = alt_fill
    r += 1

# Section 5: Status Distribution
r += 1
ws_stats.merge_cells(f'A{r}:E{r}')
ws_stats.cell(row=r, column=1, value='حالة المنافسات').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1

ws_stats.merge_cells(f'A{r}:B{r}')
ws_stats.cell(row=r, column=1, value='الحالة').font = hdr_font
ws_stats.cell(row=r, column=1).fill = hdr_fill
ws_stats.merge_cells(f'C{r}:E{r}')
ws_stats.cell(row=r, column=3, value='العدد').font = hdr_font
ws_stats.cell(row=r, column=3).fill = hdr_fill
ws_stats.cell(row=r, column=1).border = thin_border
ws_stats.cell(row=r, column=3).border = thin_border
r += 1

sorted_statuses = sorted(status_counts.items(), key=lambda x: x[1], reverse=True)
for status, count in sorted_statuses:
    ws_stats.merge_cells(f'A{r}:B{r}')
    ws_stats.cell(row=r, column=1, value=status).font = normal_font
    ws_stats.merge_cells(f'C{r}:E{r}')
    ws_stats.cell(row=r, column=3, value=f'{count:,}').font = normal_font
    ws_stats.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws_stats.cell(row=r, column=1).border = thin_border
    ws_stats.cell(row=r, column=3).border = thin_border
    if (r - 7) % 2 == 0:
        ws_stats.cell(row=r, column=1).fill = alt_fill
        ws_stats.cell(row=r, column=3).fill = alt_fill
    r += 1

ws_stats.column_dimensions['A'].width = 35
ws_stats.column_dimensions['B'].width = 10
for col in ['C','D','E']: ws_stats.column_dimensions[col].width = 18

# ──────── SHEET 8: Pricing ────────
ws_price = wb.create_sheet('عرض التسعير')
ws_price.sheet_properties.tabColor = GOLD

add_title(ws_price, 'عرض الأسعار — منتج المنافسات والمقاولين', 'Government Tenders & Contractors Database')

r = 6
ws_price.merge_cells(f'A{r}:E{r}')
ws_price.cell(row=r, column=1, value='باقات الاشتراك').font = Font(name='Arial', bold=True, size=16, color=DARK_BLUE)
ws_price.cell(row=r, column=1).alignment = Alignment(horizontal='center')
r += 2

# Pricing table
plans = [
    ('الباقة', 'المحتوى', 'التحديث', 'السعر', 'المميزات'),
    ('الباقة الأساسية\nBasic', 'قاعدة بيانات المنافسات\nلمرة واحدة', 'شهرياً', '500 ريال', 'جميع المنافسات النشطة\nفلترة حسب النشاط'),
    ('الباقة الاحترافية\nProfessional', 'المنافسات + المقاولين\nروابط التواصل', 'أسبوعياً', '1,200 ريال', 'جميع المنافسات + المقاولين\nمنافسات التشييد\nإيميلات وجوالات المقاولين'),
    ('الباقة الشاملة\nEnterprise', 'البيانات كاملة\nتقارير مخصصة', 'يومياً', '3,000 ريال', 'كل الباقات السابقة\nAPI مخصص\nتقارير حسب الطلب\nاستشارات'),
]

for i, row_data in enumerate(plans):
    for j, val in enumerate(row_data):
        cell = ws_price.cell(row=r + i, column=j + 1, value=val)
        if i == 0:
            cell.font = hdr_font
            cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            if i == 1:
                cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            elif i == 2:
                cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            elif i == 3:
                cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
            cell.font = Font(name='Arial', size=11)
            if j == 0: cell.font = Font(name='Arial', bold=True, size=11)
            if j == 3: cell.font = Font(name='Arial', bold=True, size=12, color=DARK_BLUE)

ws_price.column_dimensions['A'].width = 22
ws_price.column_dimensions['B'].width = 25
ws_price.column_dimensions['C'].width = 14
ws_price.column_dimensions['D'].width = 16
ws_price.column_dimensions['E'].width = 30

# Contact
r += len(plans) + 2
ws_price.merge_cells(f'A{r}:E{r}')
ws_price.cell(row=r, column=1, value='للطلب والاستفسار:').font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
r += 1
contact_info = [
    '📞 واتساب: +966 5X XXX XXXX',
    '📧 بريد إلكتروني: sales@example.com',
    '🌐 الموقع: example.com/tenders',
]
for line in contact_info:
    ws_price.merge_cells(f'A{r}:E{r}')
    ws_price.cell(row=r, column=1, value=line).font = Font(name='Arial', size=12)
    r += 1

r += 1
ws_price.merge_cells(f'A{r}:E{r}')
ws_price.cell(row=r, column=1, value='⚠️ ملاحظة: يتم تحديث البيانات بشكل دوري. الأسعار قابلة للتغيير.').font = Font(name='Arial', size=10, color='999999')

# ── Save ──
print(f'\nSaving Excel...')
wb.save(OUTPUT_FILE)
print(f'✓ Saved: {OUTPUT_FILE}')
print(f'  Size: {os.path.getsize(OUTPUT_FILE) / 1024 / 1024:.1f} MB')
