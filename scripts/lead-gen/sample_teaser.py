import json, os, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = 'data'
OUTPUT = os.path.join(DATA_DIR, 'عينة_مجانية_المنافسات_والمقاولين.xlsx')

def clean_str(s):
    if not s: return ''
    if not isinstance(s, str): s = str(s)
    import re
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)

with open(os.path.join(DATA_DIR, 'etimad_all_tenders.json'), 'r', encoding='utf-8') as f:
    tenders = json.load(f)
with open(os.path.join(DATA_DIR, 'muqawil_all_regions.json'), 'r', encoding='utf-8') as f:
    contractors = json.load(f)

# Pick 50 random tenders (mix of construction and others)
random.seed(42)
construction_kw = ['مقاولات','تشييد','بناء','إنشاء','هدم','ترميم','صيانة','طرق','خرسانة','دهان']
def safe_text(v): return str(v) if v is not None else ''
construction = [t for t in tenders if any(k in (safe_text(t.get('tenderName',''))+' '+safe_text(t.get('tenderActivityName',''))) for k in construction_kw)]
others = [t for t in tenders if t not in construction]
sample_tenders = random.sample(construction, min(30, len(construction))) + random.sample(others, min(20, len(others)))

# Pick 100 random contractors
sample_contractors = random.sample(contractors, min(100, len(contractors)))

wb = Workbook()

# Colors
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
title_font = Font(name='Arial', bold=True, color='1F3864', size=16)
bold_font = Font(name='Arial', bold=True, size=10)
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
gold_fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')

# Sheet 1: Cover
ws = wb.active
ws.title = 'عينة مجانية'
ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value='🎁 عينة مجانية — منتج المنافسات الحكومية والمقاولين')
c.font = Font(name='Arial', bold=True, size=20, color='1F3864')
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 50

ws.merge_cells('A3:H3')
c = ws.cell(row=3, column=1, value='هذه عينة مجانية لتقييم جودة البيانات. للاشتراك في الباقة الكاملة تواصل معنا.')
c.font = Font(name='Arial', size=11, color='666666')
c.alignment = Alignment(horizontal='center')

info = [
    ('', ''),
    ('المحتوى', 'العدد'),
    ('منافسات حكومية', '50'),
    ('مقاولين مع بيانات تواصل', '100'),
    ('', ''),
    ('للطلب والاستفسار:', ''),
    ('واتساب:', '+966 5X XXX XXXX'),
    ('بريد إلكتروني:', 'sales@example.com'),
]
for i, (k, v) in enumerate(info):
    r = 5 + i
    if k in ('المحتوى', 'للطلب والاستفسار:'):
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(row=r, column=1, value=k).font = Font(name='Arial', bold=True, size=13, color='1F3864')
    elif k and v:
        ws.merge_cells(f'A{r}:C{r}')
        ws.cell(row=r, column=1, value=k).font = bold_font
        ws.merge_cells(f'E{r}:H{r}')
        ws.cell(row=r, column=5, value=v).font = Font(name='Arial', size=11, color='2E75B6')
        ws.cell(row=r, column=5).fill = gold_fill
    elif k:
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(row=r, column=1, value=k).font = Font(name='Arial', bold=True, size=12, color='1F3864')

# Sheet 2: Sample Tenders
ws2 = wb.create_sheet('عينة المنافسات')
headers = ['اسم المنافسة', 'الجهة الحكومية', 'النشاط', 'نوع المنافسة', 'آخر موعد للتقديم', 'سعر الكراسة', 'الحالة']
for c, h in enumerate(headers, 1):
    ws2.cell(row=1, column=c, value=h).font = hdr_font
    ws2.cell(row=1, column=c).fill = hdr_fill
    ws2.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.cell(row=1, column=c).border = thin_border

status_map = {2: 'نشطة', 3: 'فتح العروض', 4: 'فحص العروض', 5: 'الترسية', 6: 'تم الترسية', 8: 'منتهية'}
for i, t in enumerate(sample_tenders, 2):
    vals = [
        t.get('tenderName', ''),
        t.get('agencyName', ''),
        t.get('tenderActivityName', ''),
        t.get('tenderTypeName', ''),
        (t.get('lastOfferPresentationDate') or '')[:10],
        t.get('condetionalBookletPrice', ''),
        status_map.get(t.get('tenderStatusId'), '')
    ]
    for j, v in enumerate(vals, 1):
        cell = ws2.cell(row=i, column=j, value=clean_str(v))
        cell.font = normal_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        if (i - 2) % 2 == 1:
            cell.fill = alt_fill

ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 14
ws2.auto_filter.ref = f'A1:G{len(sample_tenders)+1}'
ws2.freeze_panes = 'A2'

# Sheet 3: Sample Contractors
ws3 = wb.create_sheet('عينة المقاولين')
headers2 = ['اسم الشركة', 'المنطقة', 'المدينة', 'الجوال', 'البريد الإلكتروني']
for c, h in enumerate(headers2, 1):
    ws3.cell(row=1, column=c, value=h).font = hdr_font
    ws3.cell(row=1, column=c).fill = hdr_fill
    ws3.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws3.cell(row=1, column=c).border = thin_border

for i, cont in enumerate(sample_contractors, 2):
    vals = [
        cont.get('companyName', ''),
        cont.get('region_name', ''),
        cont.get('cityName', ''),
        cont.get('phone', ''),
        cont.get('email', '')
    ]
    for j, v in enumerate(vals, 1):
        cell = ws3.cell(row=i, column=j, value=clean_str(v))
        cell.font = normal_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        if (i - 2) % 2 == 1:
            cell.fill = alt_fill

ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 28
ws3.auto_filter.ref = f'A1:E{len(sample_contractors)+1}'
ws3.freeze_panes = 'A2'

wb.save(OUTPUT)
print(f'Done! Size: {os.path.getsize(OUTPUT)/1024:.0f} KB')
